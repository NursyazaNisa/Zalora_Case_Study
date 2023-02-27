import requests
from bs4 import *
import pandas as pd
import os
import openpyxl

def folder_create(images):
    RESULT_LOCATION = r"C:\Nisa\Zalora"
    DATA = r"C:\Nisa\Question 1 Dataset.xlsx"
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook['Question 1 Dataset']
    column_values = [cell.value for col in sheet.iter_cols(min_row=1, max_row=None, min_col=1, max_col=1) for cell in col]
    column_values = list(dict.fromkeys(column_values))  # removes duplicates
    try:
        for value in column_values:
            print("Creating folder: ", value)
            folderName = value
            baseDir = RESULT_LOCATION
            os.makedirs(os.path.join(baseDir, folderName))
            print("Created folder: ", folderName)
    except:
        print("Folder Exist with that name!")
        folder_create(images)

    download_images(images, folderName)

# DOWNLOAD ALL IMAGES FROM THAT URL
def download_images(images, folderName):

    # initial count is zero
    count = 0

    # print total images found in URL
    print("Total %s Image Found!") %len(images)

    # checking if images is not zero
    if len(images)!=0:
        for i, image in enumerate(images):
            # From image tag ,Fetch image Source URL

            # 1.data-srcset
            # 2.data-src
            # 3.data-fallback-src
            # 4.src

            # Here we will use exception handling

            # first we will search for "data-srcset" in img tag
            try:
                # In image tag ,searching for "data-srcset"
                image_link = image["data-srcset"]

            # then we will search for "data-src" in img
            # tag and so on..
            except:
                try:
                    # In image tag ,searching for "data-src"
                    image_link = image["data-src"]
                except:
                    try:
                        # In image tag ,searching for "data-fallback-src"
                        image_link = image["data-fallback-src"]
                    except:
                        try:
                            # In image tag ,searching for "src"
                            image_link = image["src"]

                        # if no Source URL found
                        except:
                            pass

            # After getting Image Source URL
            # We will try to get the content of image
            try:
                r = requests.get(image_link).content
                try:

                    # possibility of decode
                    r = str(r, 'utf-8')

                except UnicodeDecodeError:

                    # After checking above condition, Image Download start
                    with open("%s/images%s.jpg" % (folderName, i+1), "wb+") as f:
                        f.write(r)

                    # counting number of image downloaded
                    count += 1
            except:
                pass

        # There might be possible, that all
        # images not download
        # if all images download
        if count==len(images):
            print("All Images Downloaded!")

        # if all images not download
        else:
            print("Total %s Images Downloaded Out of %s")%count,len(images)

# --------------------------------------------------------------


# MAIN FUNCTION START
def main(url):

    DATA = r"C:\Nisa\Question 1 Dataset.xlsx"
    workbook = openpyxl.load_workbook(DATA)
    sheet = workbook['Question 1 Dataset']
    column_values = [cell.value for col in sheet.iter_cols(min_row=2, max_row=None, min_col=2, max_col=2) for cell in col]
    column_values = list(dict.fromkeys(column_values))  # removes duplicates

    for value in column_values:
        url = "https://www.zalora.com.my/catalog/?q=%s" %value
    # content of URL
        headers = {
            "User-Agent": "Chrome/110.0.5481.178",
            "Accept-Encoding": "*",
            "Connection": "keep-alive",
            "Accept": "application/json, text/javascript, */*;q=0.01",
            "Accept-Language": "en-US,en;q=0.9,fa-IR;q=0.8,fa;q=0.7,de;q=0.6,",
            "Content-Type": "application/x-www-form-urlencoded, charset=UTF-8,application/json, text/javascript"
        }
        r = requests.get(url, headers=headers)

        # Parse HTML Code
        soup = BeautifulSoup(r.text, 'html.parser')

        # find all images in URL
        images = soup.findAll('img')

        # Call folder create function
        folder_create(images)


# take url
url = 'https://www.zalora.com.my/catalog/'

# CALL MAIN FUNCTION
main(url)


